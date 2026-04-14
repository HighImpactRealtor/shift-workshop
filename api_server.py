#!/usr/bin/env python3
"""
THE SHIFT Workshop — Registration API Server

Handles:
1. Registration form submissions
2. Zoom meeting registration via API (when configured)
3. Stores registrants for reminder sequences

SETUP REQUIRED:
Set these environment variables before running:
  ZOOM_ACCOUNT_ID     — from your Server-to-Server OAuth app
  ZOOM_CLIENT_ID      — from your Server-to-Server OAuth app  
  ZOOM_CLIENT_SECRET   — from your Server-to-Server OAuth app
  ZOOM_MEETING_ID      — the meeting ID with registration enabled
  
If Zoom creds are not set, the server runs in "standalone" mode
(stores registrants locally, no Zoom API calls).
"""

import os
import io
import sqlite3
import time
import base64
from datetime import datetime, timezone
from contextlib import asynccontextmanager

import httpx
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

# ── Config ──────────────────────────────────────────────

ZOOM_ACCOUNT_ID = os.environ.get("ZOOM_ACCOUNT_ID", "")
ZOOM_CLIENT_ID = os.environ.get("ZOOM_CLIENT_ID", "")
ZOOM_CLIENT_SECRET = os.environ.get("ZOOM_CLIENT_SECRET", "")
ZOOM_MEETING_ID = os.environ.get("ZOOM_MEETING_ID", "")

ZOOM_CONFIGURED = all([ZOOM_ACCOUNT_ID, ZOOM_CLIENT_ID, ZOOM_CLIENT_SECRET, ZOOM_MEETING_ID])

# Token cache
_zoom_token = {"access_token": "", "expires_at": 0}

# ── Database ────────────────────────────────────────────

db = sqlite3.connect("registrants.db", check_same_thread=False)
db.execute("""
    CREATE TABLE IF NOT EXISTS registrants (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        first_name TEXT NOT NULL,
        last_name TEXT NOT NULL,
        email TEXT NOT NULL UNIQUE,
        phone TEXT,
        production_goal TEXT,
        stuck TEXT,
        questions TEXT,
        zoom_join_url TEXT,
        zoom_registrant_id TEXT,
        registered_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
""")
# Add new columns if they don't exist yet (for existing deployments)
for col in ["production_goal", "stuck", "questions"]:
    try:
        db.execute(f"ALTER TABLE registrants ADD COLUMN {col} TEXT")
    except sqlite3.OperationalError:
        pass  # Column already exists
db.commit()

# ── Zoom API helpers ────────────────────────────────────

async def get_zoom_token() -> str:
    """Get or refresh the Server-to-Server OAuth access token."""
    global _zoom_token
    
    if _zoom_token["access_token"] and time.time() < _zoom_token["expires_at"] - 60:
        return _zoom_token["access_token"]
    
    credentials = base64.b64encode(
        f"{ZOOM_CLIENT_ID}:{ZOOM_CLIENT_SECRET}".encode()
    ).decode()
    
    async with httpx.AsyncClient() as client:
        resp = await client.post(
            "https://zoom.us/oauth/token",
            params={
                "grant_type": "account_credentials",
                "account_id": ZOOM_ACCOUNT_ID,
            },
            headers={
                "Authorization": f"Basic {credentials}",
                "Content-Type": "application/x-www-form-urlencoded",
            },
        )
        
        if resp.status_code != 200:
            raise HTTPException(status_code=422, detail="Failed to authenticate with Zoom API")
        
        data = resp.json()
        _zoom_token["access_token"] = data["access_token"]
        _zoom_token["expires_at"] = time.time() + data.get("expires_in", 3600)
        return _zoom_token["access_token"]


async def register_with_zoom(first_name: str, last_name: str, email: str) -> dict:
    """Register a person for the Zoom meeting. Returns join_url and registrant_id."""
    token = await get_zoom_token()
    
    async with httpx.AsyncClient() as client:
        resp = await client.post(
            f"https://api.zoom.us/v2/meetings/{ZOOM_MEETING_ID}/registrants",
            headers={
                "Authorization": f"Bearer {token}",
                "Content-Type": "application/json",
            },
            json={
                "email": email,
                "first_name": first_name,
                "last_name": last_name,
            },
        )
        
        if resp.status_code == 201:
            data = resp.json()
            return {
                "join_url": data.get("join_url", ""),
                "registrant_id": data.get("registrant_id", data.get("id", "")),
            }
        elif resp.status_code == 409:
            # Already registered
            return {"join_url": "", "registrant_id": "already_registered"}
        else:
            print(f"Zoom API error: {resp.status_code} — {resp.text}")
            raise HTTPException(
                status_code=422,
                detail="Zoom registration failed. Please try again or email larkin@highimpactrealtor.com."
            )

# ── FastAPI app ─────────────────────────────────────────

@asynccontextmanager
async def lifespan(app):
    mode = "ZOOM CONNECTED" if ZOOM_CONFIGURED else "STANDALONE (no Zoom creds)"
    print(f"[SHIFT API] Running in {mode} mode")
    yield
    db.close()

app = FastAPI(lifespan=lifespan)
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

class RegistrationRequest(BaseModel):
    first_name: str
    last_name: str
    email: str
    phone: str = ""
    production_goal: str = ""
    stuck: str = ""
    questions: str = ""

@app.post("/api/register", status_code=201)
async def register(req: RegistrationRequest):
    first = req.first_name.strip()
    last = req.last_name.strip()
    email = req.email.strip().lower()
    phone = req.phone.strip()
    production_goal = req.production_goal.strip()
    stuck = req.stuck.strip()
    questions = req.questions.strip()
    
    if not first or not last or not email:
        raise HTTPException(status_code=400, detail="First name, last name, and email are required.")
    
    if "@" not in email:
        raise HTTPException(status_code=400, detail="Please enter a valid email address.")
    
    # Check for existing registration
    existing = db.execute("SELECT id FROM registrants WHERE email = ?", [email]).fetchone()
    if existing:
        return {"status": "already_registered", "message": "You're already registered. Check your inbox for your Zoom link."}
    
    zoom_join_url = ""
    zoom_registrant_id = ""
    
    # Register with Zoom if configured
    if ZOOM_CONFIGURED:
        zoom_data = await register_with_zoom(first, last, email)
        zoom_join_url = zoom_data.get("join_url", "")
        zoom_registrant_id = zoom_data.get("registrant_id", "")
    
    # Save to local DB
    try:
        db.execute(
            """INSERT INTO registrants 
               (first_name, last_name, email, phone, production_goal, stuck, questions, zoom_join_url, zoom_registrant_id) 
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            [first, last, email, phone, production_goal, stuck, questions, zoom_join_url, zoom_registrant_id]
        )
        db.commit()
    except sqlite3.IntegrityError:
        return {"status": "already_registered", "message": "You're already registered."}
    
    return {
        "status": "registered",
        "message": f"Welcome, {first}! You're registered for The SHIFT.",
        "has_zoom_link": bool(zoom_join_url),
    }

@app.get("/api/registrants")
def list_registrants():
    """Admin endpoint — list all registrants."""
    rows = db.execute(
        "SELECT id, first_name, last_name, email, phone, production_goal, stuck, questions, zoom_join_url, registered_at FROM registrants ORDER BY id DESC"
    ).fetchall()
    return [
        {
            "id": r[0],
            "first_name": r[1],
            "last_name": r[2],
            "email": r[3],
            "phone": r[4],
            "production_goal": r[5],
            "stuck": r[6],
            "questions": r[7],
            "has_zoom_link": bool(r[8]),
            "registered_at": r[9],
        }
        for r in rows
    ]

@app.get("/api/export")
def export_registrants():
    """Download all registrants as a formatted Excel file."""
    rows = db.execute(
        """SELECT first_name, last_name, email, phone, production_goal, stuck, questions, 
                  zoom_join_url, registered_at 
           FROM registrants ORDER BY registered_at ASC"""
    ).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registrants"

    # ── Header styling ──
    header_fill = PatternFill("solid", fgColor="4A7C59")   # sage green
    header_font = Font(bold=True, color="FFFFFF", size=11)
    headers = [
        "First Name", "Last Name", "Email", "Phone",
        "Production Goal (Units & Volume)", "Where They Feel Stuck",
        "Other Questions", "Zoom Join URL", "Registered At"
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 30

    # ── Data rows ──
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value or "")
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        # Alternate row shading
        if row_idx % 2 == 0:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = PatternFill("solid", fgColor="F0F5F1")

    # ── Column widths ──
    col_widths = [14, 14, 28, 16, 35, 40, 35, 45, 20]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    # ── Stream response ──
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"shift_registrants_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.get("/api/stats")
def stats():
    """Quick registration stats."""
    count = db.execute("SELECT COUNT(*) FROM registrants").fetchone()[0]
    return {"total_registrants": count, "zoom_connected": ZOOM_CONFIGURED}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
